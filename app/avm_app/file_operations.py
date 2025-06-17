import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import io
import numpy as np
import seaborn as sns

from avm_app.avm_utils import get_keyword_from_model, find_file_with_keyword

def read_files_once(unique_models, avm_folder):
    """
    Reads CSV/XLSX files only once per model, using a keyword
    extracted from the model name, and returns a dict of {model: dataframe}.
    """
    model_file_data = {}
    for model in unique_models:
        keyword = get_keyword_from_model(model)
        if keyword:
            file_name = find_file_with_keyword(avm_folder, keyword)
            if file_name:
                file_path = os.path.join(avm_folder, file_name)
                # Read either CSV or Excel
                if file_path.endswith('.csv'):
                    model_df = pd.read_csv(file_path)
                elif file_path.endswith('.xlsx'):
                    model_df = pd.read_excel(file_path)
                else:
                    # Skip unsupported formats
                    continue

                # Filter to rows that match the AVM Model Name (for some models)
                if model in ['SiteXValue', 'RVM', 'ValueSure']:
                    model_df = model_df[model_df['AVM Model Name'] == model]

                model_file_data[model] = model_df
    return model_file_data

def autosize_columns(worksheet):
    """
    Auto-size the columns in an openpyxl worksheet based on max text length.
    """
    for col in worksheet.columns:
        max_length = 0
        column_letter = col[0].column_letter  # e.g. 'A', 'B', ...
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Add a little extra space
        worksheet.column_dimensions[column_letter].width = max_length + 2

def add_histograms_to_excel(file_path, output_file_path):
    """
    Generates KDE histograms (including Model 1/2/3 breakdown) for
    '% Diff between AVM and Benchmark' from the 'Original Data' sheet.
    Inserts a single combined image into a "KDE Curves" sheet.
    """
    sheet_name = "Original Data"

    # Read the specific sheet from the updated workbook
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Multiply the differences by 100 to treat them as integer-like percentages
    # If you already multiply them earlier in your code, remove or adapt this step.
    if '% Diff between AVM and Benchmark' in df.columns:
        df["% Diff between AVM and Benchmark"] = df["% Diff between AVM and Benchmark"] * 100

        # Make sure the column is numeric
        df["% Diff between AVM and Benchmark"] = pd.to_numeric(df["% Diff between AVM and Benchmark"], errors='coerce')

        # Drop NaNs and non-finite values
        data = df["% Diff between AVM and Benchmark"].dropna()
        data = data[np.isfinite(data)]
    else:
        # If the column doesn't exist, skip plotting
        data = pd.Series([])

    # Filter data by 'Model 1', 'Model 2', 'Model 3'
    models = ['Model 1', 'Model 2', 'Model 3']
    data_models = {}
    for model in models:
        if 'Model Position' in df.columns:
            subset = df[df['Model Position'] == model].get("% Diff between AVM and Benchmark", pd.Series([]))
            subset = pd.to_numeric(subset.dropna(), errors='coerce')
            data_models[model] = subset[np.isfinite(subset)]
        else:
            data_models[model] = pd.Series([])

    # Load the workbook we just wrote
    workbook = load_workbook(file_path)

    # Create a figure with multiple subplots
    plt.figure(figsize=(14, 10))

    def add_stats(ax, data_array):
        """
        Annotate each subplot with mean, median, std dev, or a no-data message.
        """
        if not data_array.empty:
            mean = data_array.mean()
            median = data_array.median()
            std_dev = data_array.std()
            ax.annotate(
                f'Mean: {mean:.4f}%\nMedian: {median:.4f}%\nStd Dev: {std_dev:.4f}%',
                xy=(0.05, 0.95), xycoords='axes fraction',
                verticalalignment='top', fontsize=9,
                bbox=dict(
                    boxstyle="round,pad=0.3",
                    edgecolor='black',
                    facecolor='white'
                )
            )
        else:
            ax.annotate(
                'No valid data available',
                xy=(0.5, 0.5), xycoords='axes fraction',
                verticalalignment='center',
                horizontalalignment='center',
                fontsize=9
            )

    # 1) KDE for all data (if any)
    ax1 = plt.subplot(2, 2, 1)
    if not data.empty:
        sns.kdeplot(data, ax=ax1, fill=True, bw_adjust=0.5, clip=(-100, 100))
    ax1.set_title('KDE (All Models)')
    ax1.set_xlabel('% Diff between AVM and Benchmark')
    ax1.set_ylabel('Density')
    add_stats(ax1, data)

    # 2) KDE for each model
    subplot_index = 2
    for model in models:
        ax = plt.subplot(2, 2, subplot_index)
        model_data = data_models[model]
        if not model_data.empty:
            sns.kdeplot(model_data, ax=ax, fill=True, bw_adjust=0.5, clip=(-100, 100))
        ax.set_title(f'KDE ({model})')
        ax.set_xlabel('% Diff between AVM and Benchmark')
        ax.set_ylabel('Density')
        add_stats(ax, model_data)
        subplot_index += 1

    plt.tight_layout()

    # Save plot to a buffer
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', bbox_inches='tight')
    buffer.seek(0)
    plt.close()

    # Insert image into a new or existing "KDE Curves" sheet
    new_sheet_name = "KDE Curves"
    if new_sheet_name not in workbook.sheetnames:
        new_sheet = workbook.create_sheet(new_sheet_name)
    else:
        new_sheet = workbook[new_sheet_name]

    img = Image(buffer)
    new_sheet.add_image(img, 'A1')

    # Save the updated workbook
    workbook.save(output_file_path)
    print("KDE curves with statistical annotations added to the Excel workbook.")

def write_results_to_excel(results_df, output_file, min_conf_scores, max_fsd_values):
    """
    Writes the main DataFrame and various calculated statistics to 'output_file'.

    -- In this version, we compute PPE10 manually by referencing the
       'Benchmark Value' and 'AVM Value' columns, ignoring any precomputed
       '% Diff between AVM and Benchmark'.

       We produce only ONE sheet for PPE10 stats: "PPE10 Stats",
       which shows total # of valid rows, how many are within 10%, and
       the resulting proportion.

    -- The rest of the sheets remain as in your original approach.
    """
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # ----------------------------------------------------------------------
        # 1. Original Data
        # ----------------------------------------------------------------------
        results_df.to_excel(writer, sheet_name='Original Data', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Original Data']

        # Optionally format '% Diff between AVM and Benchmark' if it exists
        if '% Diff between AVM and Benchmark' in results_df.columns:
            percent_col_idx = results_df.columns.get_loc('% Diff between AVM and Benchmark') + 1
            for row in range(2, len(results_df) + 2):
                worksheet.cell(row=row, column=percent_col_idx).number_format = '0.00%'

        autosize_columns(worksheet)

        # ----------------------------------------------------------------------
        # 2. Model-Specific Statistics (unchanged)
        # ----------------------------------------------------------------------
        if '% Diff between AVM and Benchmark' in results_df.columns:
            overall_avg_error = results_df['% Diff between AVM and Benchmark'].mean(skipna=True)
            model_stats_df = (
                results_df
                .groupby('Model Position')['% Diff between AVM and Benchmark']
                .agg(['mean', 'median', 'std'])
                .reset_index()
            )
            model_stats_df.columns = ['Model Position', 'Average Error', 'Median Error', 'Standard Deviation']
            model_stats_df[['Average Error','Median Error','Standard Deviation']] = \
                model_stats_df[['Average Error','Median Error','Standard Deviation']].apply(pd.to_numeric, errors='coerce').round(3)

            # Add an "Overall" row
            overall_stats = pd.DataFrame([[
                'Overall',
                round(overall_avg_error, 3),
                round(results_df['% Diff between AVM and Benchmark'].median(), 3),
                round(results_df['% Diff between AVM and Benchmark'].std(), 3)
            ]], columns=['Model Position', 'Average Error', 'Median Error', 'Standard Deviation'])

            model_stats_df = pd.concat([model_stats_df, overall_stats], ignore_index=True)
            model_stats_df.to_excel(writer, sheet_name='Model Specific Statistics', index=False)

            stats_worksheet = writer.sheets['Model Specific Statistics']
            for col_name in ['Average Error', 'Median Error', 'Standard Deviation']:
                col_idx = model_stats_df.columns.get_loc(col_name) + 1
                for row in range(2, len(model_stats_df) + 2):
                    stats_worksheet.cell(row=row, column=col_idx).number_format = '0.00%'

            autosize_columns(stats_worksheet)

        # ----------------------------------------------------------------------
        # 3. Statistics by Location (% Diff) [Optional - unchanged]
        # ----------------------------------------------------------------------
        if 'AVM Value' in results_df.columns and '% Diff between AVM and Benchmark' in results_df.columns:
            filtered_results_df = results_df[results_df['AVM Value'].notna()].copy()
            loc_stats = (
                filtered_results_df
                .groupby(['State', 'County'])['% Diff between AVM and Benchmark']
                .agg(['count','mean','min','max'])
                .reset_index()
            )
            expected_cols = ['State','County','Hits','Average Error','Minimum Error','Maximum Error']
            if loc_stats.shape[1] == len(expected_cols):
                loc_stats.columns = expected_cols
            else:
                print("Unexpected column count in loc_stats:", loc_stats.shape[1])
                print("Columns were:", loc_stats.columns)
                raise ValueError("Mismatch between actual and expected column count when renaming.")


            # Calculate total records for each (State, County)
            total_records = (
                results_df
                .groupby(['State','County'])
                .size()
                .reset_index(name='Total Number of Records')
            )
            loc_stats = pd.merge(total_records, loc_stats, on=['State','County'], how='left')

            loc_stats['Hit Rate'] = loc_stats['Hits'] / loc_stats['Total Number of Records']
            loc_stats['Hit Rate'] = loc_stats['Hit Rate'].round(3)
            print("loc_stats type:", type(loc_stats))
            print("loc_stats columns:", loc_stats.columns)
            print("loc_stats head:\n", loc_stats.head())
            print("loc_stats['Average Error'] type:", type(loc_stats['Average Error']))

            if isinstance(loc_stats['Average Error'], pd.Series):
                loc_stats['Average Error'] = pd.to_numeric(loc_stats['Average Error'], errors='coerce').round(3)
            else:
                loc_stats['Average Error'] = round(float(loc_stats['Average Error']), 3)

            loc_stats['Minimum Error'] = pd.to_numeric(loc_stats['Minimum Error'], errors='coerce').round(3)
            loc_stats['Maximum Error'] = pd.to_numeric(loc_stats['Maximum Error'], errors='coerce').round(3)
            

            loc_stats = loc_stats[
                ['State','County','Total Number of Records','Hits','Hit Rate','Average Error','Minimum Error','Maximum Error']
            ]
            loc_stats.to_excel(writer, sheet_name='Statistics by Location', index=False)

            stats_location_worksheet = writer.sheets['Statistics by Location']
            # Format relevant columns as percentages
            hit_rate_col = loc_stats.columns.get_loc('Hit Rate') + 1
            avg_err_col = loc_stats.columns.get_loc('Average Error') + 1
            min_err_col = loc_stats.columns.get_loc('Minimum Error') + 1
            max_err_col = loc_stats.columns.get_loc('Maximum Error') + 1

            for row in range(2, len(loc_stats) + 2):
                stats_location_worksheet.cell(row=row, column=hit_rate_col).number_format = '0.00%'
                stats_location_worksheet.cell(row=row, column=avg_err_col).number_format = '0.00%'
                stats_location_worksheet.cell(row=row, column=min_err_col).number_format = '0.00%'
                stats_location_worksheet.cell(row=row, column=max_err_col).number_format = '0.00%'

            autosize_columns(stats_location_worksheet)

        # ----------------------------------------------------------------------
        # 4. Model Usage (unchanged)
        # ----------------------------------------------------------------------
        if 'Model Position' in results_df.columns:
            model_usage = results_df['Model Position'].value_counts(normalize=True)
            model_usage_df = model_usage.reset_index()
            model_usage_df.columns = ['Model Position', 'Usage Percentage']
            model_usage_df.to_excel(writer, sheet_name='Model Usage', index=False)

            usage_worksheet = writer.sheets['Model Usage']
            usage_percentage_col_idx = model_usage_df.columns.get_loc('Usage Percentage') + 1
            for row in range(2, len(model_usage_df) + 2):
                usage_worksheet.cell(row=row, column=usage_percentage_col_idx).number_format = '0.0%'

            autosize_columns(usage_worksheet)

        # ----------------------------------------------------------------------
        # 5. Model Name Counts (unchanged)
        # ----------------------------------------------------------------------
        if 'AVM Name' in results_df.columns:
            model_name_counts = results_df['AVM Name'].value_counts()
            model_name_counts_df = model_name_counts.reset_index()
            model_name_counts_df.columns = ['AVM Name', 'Count']
            model_name_counts_df.to_excel(writer, sheet_name='Model Name Counts', index=False)

            autosize_columns(writer.sheets['Model Name Counts'])

        # ----------------------------------------------------------------------
        # 6. Conf & FSD Summary (unchanged)
        # ----------------------------------------------------------------------
        conf_fsd_summary_df = pd.DataFrame({
            'Model Name': min_conf_scores.keys(),
            'Min Conf Score': min_conf_scores.values(),
            'Max FSD Value': max_fsd_values.values()
        })
        conf_fsd_summary_df.to_excel(writer, sheet_name='Conf & FSD Summary', index=False)
        autosize_columns(writer.sheets['Conf & FSD Summary'])

        # ----------------------------------------------------------------------
        # 7. Manual PPE10 Stats (ONE sheet)
        # ----------------------------------------------------------------------
        if 'Benchmark Value' in results_df.columns and 'AVM Value' in results_df.columns:
            df_valid = results_df[
                results_df['Benchmark Value'].notna() &
                results_df['AVM Value'].notna() &
                (results_df['Benchmark Value'] != 0)  # avoid division by zero
            ].copy()

            # Compute absolute percentage difference
            df_valid['abs_pct_diff'] = (
                (df_valid['AVM Value'] - df_valid['Benchmark Value']).abs()
                / df_valid['Benchmark Value'].abs()
            ) * 100.0

            # Mark within ±10%
            df_valid['Within_10'] = df_valid['abs_pct_diff'] <= 10

            total_records = len(df_valid)
            count_within_10 = df_valid['Within_10'].sum()
            ppe10 = count_within_10 / total_records if total_records else 0

            ppe10_df = pd.DataFrame([{
                'Total Records': total_records,
                'Count Within 10%': count_within_10,
                'PPE10': round(ppe10, 3)
            }])

            sheet_name = "PPE10 Stats"
            ppe10_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ppe10_ws = writer.sheets[sheet_name]

            # Format the "PPE10" column as 0.00%
            ppe10_col_idx = ppe10_df.columns.get_loc('PPE10') + 1
            for row_idx in range(2, len(ppe10_df) + 2):
                ppe10_ws.cell(row=row_idx, column=ppe10_col_idx).number_format = '0.00%'

            autosize_columns(ppe10_ws)
        else:
            print("Either 'Benchmark Value' or 'AVM Value' is missing; cannot compute manual PPE10.")
        # Keep all original sheets intact and add PPE10 breakdowns
        if 'Benchmark Value' in results_df.columns and 'AVM Value' in results_df.columns:
            df_valid = results_df.dropna(subset=['Benchmark Value', 'AVM Value'])
            df_valid = df_valid[df_valid['Benchmark Value'] != 0]

            df_valid['abs_pct_diff'] = (
                (df_valid['AVM Value'] - df_valid['Benchmark Value']).abs()
                / df_valid['Benchmark Value'].abs()
            ) * 100.0
            df_valid['Within_10'] = df_valid['abs_pct_diff'] <= 10

            county_ppe10 = df_valid.groupby(['State', 'County'])['Within_10'].agg(['count', 'sum'])
            county_ppe10['PPE10'] = county_ppe10['sum'] / county_ppe10['count']
            county_ppe10.to_excel(writer, sheet_name='PPE10 by County')
            autosize_columns(writer.sheets['PPE10 by County'])

            state_ppe10 = df_valid.groupby('State')['Within_10'].agg(['count', 'sum'])
            state_ppe10['PPE10'] = state_ppe10['sum'] / state_ppe10['count']
            state_ppe10.to_excel(writer, sheet_name='PPE10 by State')
            autosize_columns(writer.sheets['PPE10 by State'])

            model_ppe10 = df_valid.groupby('Model Position')['Within_10'].agg(['count', 'sum'])
            model_ppe10['PPE10'] = model_ppe10['sum'] / model_ppe10['count']
            model_ppe10.to_excel(writer, sheet_name='PPE10 by Model')
            autosize_columns(writer.sheets['PPE10 by Model'])

    # --------------------------------------------------------------------------
    # Add histograms to the final output file
    # --------------------------------------------------------------------------
    add_histograms_to_excel(output_file, output_file)
